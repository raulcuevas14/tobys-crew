"""
Toby's Crew — Investigador de Precios
Busca precios en Google Shopping (SerpApi) + Mercado Libre (API oficial)
y genera reporte Excel comparativo.
"""

import json, re, statistics, sys, time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── CONFIGURACIÓN ─────────────────────────────────────────────────────────────
SERPAPI_KEY  = "644f5c0c33dee2cd3c610fba9a5c1870fe2896bf3f400f1fc8d1816dafefd77b"
RESULTADOS   = 10
PAIS_SERP    = "mx"
SITE_ID_ML   = "MLM"
OUTPUT_DIR   = Path("reportes")
# ──────────────────────────────────────────────────────────────────────────────

def cargar_productos(fuente):
    p = Path(fuente)
    if not p.exists():
        print(f"❌ No se encontró: {fuente}"); sys.exit(1)
    if p.suffix == ".json":
        data = json.loads(p.read_text(encoding="utf-8"))
        return [{"nombre": f"{x.get('brand','')} {x.get('name','')}".strip(),
                 "peso": x.get("weight",""),
                 "query": f"{x.get('brand','')} {x.get('name','')} {x.get('weight','')} alimento mascota".strip(),
                 "precio_actual": x.get("price", 0)} for x in data]
    elif p.suffix == ".txt":
        return [{"nombre": l.strip(), "peso": "", "query": f"{l.strip()} alimento mascota", "precio_actual": 0}
                for l in p.read_text(encoding="utf-8").splitlines()
                if l.strip() and not l.startswith("#")]
    print("❌ Usa .json o .txt"); sys.exit(1)

def buscar_google_shopping(query):
    try:
        res = requests.get("https://serpapi.com/search", timeout=15, params={
            "engine": "google_shopping", "q": query, "gl": PAIS_SERP,
            "hl": "es", "currency": "MXN", "num": RESULTADOS, "api_key": SERPAPI_KEY})
        res.raise_for_status()
        out = []
        for item in res.json().get("shopping_results", [])[:RESULTADOS]:
            precio = limpiar_precio(item.get("price",""))
            if precio:
                out.append({"fuente":"Google Shopping","tienda":item.get("source","?"),
                            "nombre":item.get("title",""),"precio":precio,"link":item.get("link","")})
        return out
    except Exception as e:
        print(f"    ⚠️  GS error: {e}"); return []

def buscar_mercado_libre(query):
    try:
        res = requests.get(f"https://api.mercadolibre.com/sites/{SITE_ID_ML}/search",
                           timeout=15, params={"q": query, "limit": RESULTADOS})
        res.raise_for_status()
        out = []
        for item in res.json().get("results", [])[:RESULTADOS]:
            precio = item.get("price")
            if precio:
                out.append({"fuente":"Mercado Libre",
                            "tienda": item.get("seller",{}).get("nickname","Vendedor ML"),
                            "nombre": item.get("title",""), "precio": float(precio),
                            "link":   item.get("permalink","")})
        return out
    except Exception as e:
        print(f"    ⚠️  ML error: {e}"); return []

def limpiar_precio(texto):
    if not texto: return None
    try: return float(re.sub(r"[^\d.,]","",str(texto)).replace(",",""))
    except: return None

def calcular_metricas(precios):
    if not precios: return {}
    n = len(precios); prom = statistics.mean(precios)
    desv = statistics.stdev(precios) if n > 1 else 0
    ps   = sorted(precios)
    return {
        "n": n, "promedio": round(prom,2), "mediana": round(statistics.median(precios),2),
        "minimo": round(min(precios),2), "maximo": round(max(precios),2),
        "desv_std": round(desv,2), "cv": round(desv/prom*100,1) if prom else 0,
        "iqr": round(ps[min(n-1,int(n*.75))] - ps[max(0,int(n*.25)-1)],2),
        "precio_p40": round(ps[max(0,int(n*.40)-1)],2),
    }

# ── ESTILOS ───────────────────────────────────────────────────────────────────
C_DARK="1A1208"; C_GOLD="C8860A"; C_GREEN="2D6A4F"
C_GREEN_LT="D8F0E5"; C_RED_LT="FDE8E8"; C_ML="FFE600"; C_GRAY="F7F5F0"; C_WHITE="FFFFFF"

def hf(color=C_WHITE,sz=9): return Font(name="Arial",bold=True,color=color,size=sz)
def cf(bold=False,color="000000",sz=9): return Font(name="Arial",bold=bold,color=color,size=sz)
def fl(color): return PatternFill("solid",fgColor=color)
def bd():
    s=Side(style="thin",color="DDDDDD"); return Border(left=s,right=s,top=s,bottom=s)
def ct(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def lf(): return Alignment(horizontal="left",vertical="center",wrap_text=True)

def crear_hoja_resumen(wb, todos):
    ws = wb.active; ws.title="Resumen"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:P1")
    ws["A1"]="🐾  Toby's Crew — Investigación de Precios  |  Google Shopping + Mercado Libre México"
    ws["A1"].font=Font(name="Arial",bold=True,size=13,color=C_WHITE)
    ws["A1"].fill=fl(C_DARK); ws["A1"].alignment=ct(); ws.row_dimensions[1].height=34

    ws.merge_cells("A2:P2")
    ws["A2"]=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Fuentes: Google Shopping + Mercado Libre API"
    ws["A2"].font=Font(name="Arial",size=8,color="888888"); ws["A2"].fill=fl("F0EDE8")
    ws["A2"].alignment=ct(); ws.row_dimensions[2].height=16

    grupos=[("",1,3),("GOOGLE SHOPPING",4,8),("MERCADO LIBRE",9,13),("COMBINADO",14,16)]
    ws.row_dimensions[3].height=18
    for label,ci,cf_ in grupos:
        if ci==cf_: cell=ws.cell(row=3,column=ci,value=label)
        else:
            ws.merge_cells(start_row=3,start_column=ci,end_row=3,end_column=cf_)
            cell=ws.cell(row=3,column=ci,value=label)
        bg={"GOOGLE SHOPPING":"1A73E8","MERCADO LIBRE":C_ML,"COMBINADO":C_GREEN}.get(label,C_DARK)
        fc=C_DARK if label=="MERCADO LIBRE" else C_WHITE
        cell.font=Font(name="Arial",bold=True,size=9,color=fc); cell.fill=fl(bg)
        cell.alignment=ct(); cell.border=bd()

    hdrs=["Producto","Peso","Tu Precio\nActual",
          "n GS","Mín GS","Prom GS","Med GS","P40 GS",
          "n ML","Mín ML","Prom ML","Med ML","P40 ML",
          "Mín\nGlobal","P40\nSugerido","Diferencia\nvs Actual"]
    ws.row_dimensions[4].height=36
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=4,column=col,value=h)
        c.font=hf(); c.fill=fl(C_DARK); c.alignment=ct(); c.border=bd()

    money_cols={3,5,6,7,8,10,11,12,13,14,15}
    for i,prod in enumerate(todos):
        row=i+5; bg=C_WHITE if i%2==0 else C_GRAY
        mgs=prod.get("metricas_gs",{}); mml=prod.get("metricas_ml",{}); mall=prod.get("metricas_all",{})
        pa=prod.get("precio_actual") or 0; p40=mall.get("precio_p40") or 0
        diff=round(p40-pa,2) if pa and p40 else ""
        vals=[prod["nombre"],prod.get("peso",""),pa or "",
              mgs.get("n",0),mgs.get("minimo",""),mgs.get("promedio",""),mgs.get("mediana",""),mgs.get("precio_p40",""),
              mml.get("n",0),mml.get("minimo",""),mml.get("promedio",""),mml.get("mediana",""),mml.get("precio_p40",""),
              mall.get("minimo",""),mall.get("precio_p40",""),diff]
        ws.row_dimensions[row].height=20
        for col,val in enumerate(vals,1):
            c=ws.cell(row=row,column=col,value=val)
            c.font=cf(); c.fill=fl(bg); c.border=bd()
            c.alignment=ct() if col>2 else lf()
            if col in money_cols and val: c.number_format='"$"#,##0.00'
        dc=ws.cell(row=row,column=16)
        if diff!="":
            dc.number_format='"$"#,##0.00;[Red]("$"#,##0.00)'
            dc.font=cf(bold=True,color=C_GREEN if diff>=0 else "A32D2D")
            dc.fill=fl(C_GREEN_LT if diff>=0 else C_RED_LT)
        dc.border=bd(); dc.alignment=ct()

    anchos=[30,7,11,9,10,10,10,10,9,10,10,10,10,10,11,12]
    for col,w in enumerate(anchos,1): ws.column_dimensions[get_column_letter(col)].width=w
    lr=len(todos)+6; ws.merge_cells(f"A{lr}:P{lr}")
    ws[f"A{lr}"]="📊  P40 = percentil 40 (competitivo sin ser el más barato)  ·  Verde = puedes subir precio  ·  Rojo = estás por encima del mercado"
    ws[f"A{lr}"].font=Font(name="Arial",size=8,italic=True,color="888888"); ws[f"A{lr}"].alignment=lf()

def crear_hoja_detalle(wb, prod):
    ws=wb.create_sheet(title=prod["nombre"][:28].replace("/","-"))
    ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1")
    ws["A1"]=f"🐾  {prod['nombre']}  {prod.get('peso','')}"
    ws["A1"].font=Font(name="Arial",bold=True,size=12,color=C_WHITE)
    ws["A1"].fill=fl(C_DARK); ws["A1"].alignment=ct(); ws.row_dimensions[1].height=28

    fila=2
    for label,m,color,fc in [("Google Shopping",prod.get("metricas_gs",{}),"1A73E8",C_WHITE),
                               ("Mercado Libre",prod.get("metricas_ml",{}),C_ML,C_DARK),
                               ("Combinado",prod.get("metricas_all",{}),C_GREEN,C_WHITE)]:
        ws.merge_cells(f"A{fila}:F{fila}")
        ws[f"A{fila}"]=f"{label}  ·  Mín:${m.get('minimo','—')}  Prom:${m.get('promedio','—')}  Med:${m.get('mediana','—')}  P40:${m.get('precio_p40','—')}  n={m.get('n',0)}"
        ws[f"A{fila}"].font=Font(name="Arial",size=9,bold=True,color=fc)
        ws[f"A{fila}"].fill=fl(color); ws[f"A{fila}"].alignment=lf()
        ws.row_dimensions[fila].height=18; fila+=1

    hdrs=["#","Fuente","Tienda","Producto","Precio","Link"]
    ws.row_dimensions[fila].height=20
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=fila,column=col,value=h)
        c.font=hf(); c.fill=fl(C_DARK); c.alignment=ct(); c.border=bd()
    fila+=1

    todos_r=sorted(prod.get("resultados_gs",[])+prod.get("resultados_ml",[]),key=lambda x:x["precio"])
    for i,r in enumerate(todos_r):
        bg=C_WHITE if i%2==0 else C_GRAY
        ws.row_dimensions[fila].height=16
        for col,val in enumerate([i+1,r["fuente"],r["tienda"],r["nombre"],r["precio"],r.get("link","")],1):
            c=ws.cell(row=fila,column=col,value=val)
            c.font=cf(sz=9); c.fill=fl("E8F0FE" if col==2 and r["fuente"]=="Google Shopping" else C_ML if col==2 else bg)
            c.border=bd(); c.alignment=ct() if col in(1,5) else lf()
            if col==5: c.number_format='"$"#,##0.00'
        fila+=1

    if not todos_r:
        ws.merge_cells(f"A{fila}:F{fila}")
        ws[f"A{fila}"]="Sin resultados"; ws[f"A{fila}"].font=Font(name="Arial",size=10,italic=True,color="888888")
        ws[f"A{fila}"].alignment=ct()

    for col,w in enumerate([4,16,22,40,14,50],1):
        ws.column_dimensions[get_column_letter(col)].width=w

def generar_excel(todos):
    OUTPUT_DIR.mkdir(exist_ok=True)
    fp=OUTPUT_DIR/f"precios_mercado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb=Workbook(); crear_hoja_resumen(wb,todos)
    for p in todos: crear_hoja_detalle(wb,p)
    wb.save(fp); return fp

def main():
    if len(sys.argv)<2:
        print("\nUso:\n  python3 buscar_precios.py products.json\n  python3 buscar_precios.py productos.txt\n"); sys.exit(1)
    if SERPAPI_KEY=="TU_API_KEY_AQUI":
        print("❌ Configura tu SERPAPI_KEY primero.\n"); sys.exit(1)

    productos=cargar_productos(sys.argv[1])
    print(f"\n🐾  Toby's Crew — Investigador de Precios")
    print(f"{'─'*55}")
    print(f"📦  Productos: {len(productos)}  |  Google Shopping + Mercado Libre\n")

    todos=[]
    for i,prod in enumerate(productos,1):
        print(f"[{i}/{len(productos)}] {prod['nombre']}")
        print("   🔍 Google Shopping..."); res_gs=buscar_google_shopping(prod["query"]); time.sleep(0.5)
        print("   🛒 Mercado Libre...");   res_ml=buscar_mercado_libre(prod["query"])
        mgs=calcular_metricas([r["precio"] for r in res_gs])
        mml=calcular_metricas([r["precio"] for r in res_ml])
        mall=calcular_metricas([r["precio"] for r in res_gs+res_ml])
        prod.update({"resultados_gs":res_gs,"resultados_ml":res_ml,
                     "metricas_gs":mgs,"metricas_ml":mml,"metricas_all":mall})
        todos.append(prod)
        print(f"   ✅ GS: ${mgs.get('precio_p40','—')}  |  ML: ${mml.get('precio_p40','—')}  |  Sugerido: ${mall.get('precio_p40','—')}\n")

    print("📊 Generando Excel...")
    archivo=generar_excel(todos)
    print(f"✅ Guardado en: {archivo}\n")

if __name__=="__main__":
    main()
